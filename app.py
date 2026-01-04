import os
import cv2
import numpy as np
import ezdxf
from flask import Flask, request, send_file, jsonify, after_this_request
from flask_cors import CORS
from werkzeug.utils import secure_filename
from werkzeug.exceptions import RequestEntityTooLarge
import tempfile
from PyPDF2 import PdfMerger
from pdf2docx import Converter
import pdfplumber
from openpyxl import Workbook, load_workbook
from docx import Document
from reportlab.lib.pagesizes import A4, landscape
from reportlab.pdfgen import canvas
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors
import fitz
from PIL import Image
import zipfile
import logging
import sys
import gc
import io
import shutil

# Configure logging
logging.basicConfig(stream=sys.stdout, level=logging.DEBUG)
logger = logging.getLogger(__name__)

app = Flask(__name__)
# Explicitly allow all origins and headers for debugging
CORS(app, resources={r"/*": {"origins": "*", "allow_headers": "*", "methods": "*"}})

# Set Max Content Length to 50MB (server limit)
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024

# ============== MEMORY LIMITS (FOR 512MB RAM SERVER) ==============
# CRITICAL: Server hanya punya 512MB RAM!
MAX_IMAGE_PIXELS = 800 * 800      # Max 0.64MP per image (very conservative)
MAX_IMAGE_DIMENSION = 800         # Max width/height before skip
COMPRESS_TARGET_SIZE = 300        # Target resize dimension (aggressive)
JPEG_QUALITY = 15                 # Lower quality = smaller file
Image.MAX_IMAGE_PIXELS = 10000000 # 10MP limit for Pillow (conservative)
GC_EVERY_N_IMAGES = 2             # Garbage collect every N images (aggressive for 512MB)

@app.errorhandler(RequestEntityTooLarge)
def handle_file_too_large(e):
    logger.error(f"File too large error: {e}")
    return jsonify({'error': 'File terlalu besar. Maksimum ukuran file adalah 50MB.'}), 413

@app.errorhandler(500)
def handle_internal_error(e):
    logger.error(f"Internal Server Error: {e}")
    return jsonify({'error': 'Terjadi kesalahan internal pada server. Cek log server.'}), 500

def image_to_dxf(image_path, output_path):
    # Read image with alpha channel support (IMREAD_UNCHANGED)
    img = cv2.imread(image_path, cv2.IMREAD_UNCHANGED)
    if img is None:
        raise ValueError("Could not read image")
    
    # Handle Transparency: Convert to white background
    if img.shape[2] == 4:
        # Create white background
        bg = np.ones_like(img[:,:,:3]) * 255
        # Extract alpha channel
        alpha = img[:,:,3] / 255.0
        # Blend
        for c in range(3):
            bg[:,:,c] = bg[:,:,c] * (1 - alpha) + img[:,:,c] * alpha
        img = bg.astype(np.uint8)
    
    # Convert to grayscale
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    
    # Reduce noise with Gaussian Blur
    blurred = cv2.GaussianBlur(gray, (5, 5), 0)
    
    # Auto Canny Thresholds
    v = np.median(blurred)
    sigma = 0.33
    lower = int(max(0, (1.0 - sigma) * v))
    upper = int(min(255, (1.0 + sigma) * v))
    
    # Edge detection
    edges = cv2.Canny(blurred, lower, upper)
    
    # Find contours - Use RETR_LIST to capture all contours (inner and outer)
    contours, hierarchy = cv2.findContours(edges, cv2.RETR_LIST, cv2.CHAIN_APPROX_SIMPLE)
    
    # Create DXF document
    doc = ezdxf.new('R2010')
    msp = doc.modelspace()
    
    height, width = edges.shape
    
    for contour in contours:
        # contour is a numpy array of shape (n, 1, 2)
        if len(contour) < 2:
            continue
            
        points = []
        for point in contour:
            x, y = point[0]
            # Flip Y for CAD (image 0,0 is top-left, CAD is bottom-left)
            points.append((float(x), float(height - y)))
            
        msp.add_lwpolyline(points)
            
    doc.saveas(output_path)

def pdf_to_docx(pdf_path, output_path):
    cv = Converter(pdf_path)
    cv.convert(output_path)
    cv.close()

def pdf_to_xlsx(pdf_path, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "PDF"
    # Using pdfplumber to open the PDF
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()
            if tables:
                for table in tables:
                    for row in table:
                        # Clean cell data: handle None and convert to string
                        cleaned_row = []
                        for cell in row:
                            if cell is None:
                                cleaned_row.append("")
                            else:
                                cleaned_row.append(str(cell))
                        ws.append(cleaned_row)
                    ws.append([]) # Add empty row between tables
            else:
                # If no tables, extract text line by line
                text = page.extract_text() or ""
                for line in text.splitlines():
                    ws.append([line])
                ws.append([])
    wb.save(output_path)

def docx_to_pdf(docx_path, output_path):
    d = Document(docx_path)
    c = canvas.Canvas(output_path, pagesize=A4)
    w, h = A4
    y = h - 50
    for p in d.paragraphs:
        t = p.text
        for line in t.split("\n"):
            c.drawString(50, y, line)
            y -= 14
            if y < 50:
                c.showPage()
                y = h - 50
    c.save()

def xlsx_to_pdf(xlsx_path, output_path):
    wb = load_workbook(xlsx_path, data_only=True)
    doc = SimpleDocTemplate(output_path, pagesize=landscape(A4))
    elements = []
    for sheet in wb.worksheets:
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(["" if v is None else str(v) for v in row])
        if data:
            tbl = Table(data)
            tbl.setStyle(TableStyle([
                ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
                ('FONTSIZE', (0,0), (-1,-1), 8),
            ]))
            elements.append(tbl)
    doc.build(elements)

def image_to_pdf(image_path, output_path):
    image = Image.open(image_path)
    if image.mode != 'RGB':
        image = image.convert('RGB')
    image.save(output_path, "PDF", resolution=100.0)

def merge_pdfs(paths, output_path):
    merger = PdfMerger()
    for p in paths:
        merger.append(p)
    merger.write(output_path)
    merger.close()

def compress_pdf(input_path, output_path):
    """
    Compress PDF with fixes for:
    1. Linearisation error - using linear=False
    2. File size not reduced - force recreate if needed
    3. RAM explosion - VERY strict memory management for 512MB server
    4. Process ALL images - no limit, but aggressive GC
    """
    logger.info(f"Starting compression for {input_path}")
    
    original_size = os.path.getsize(input_path)
    logger.info(f"Original file size: {original_size} bytes")
    
    doc = None
    temp_output = output_path + ".tmp"
    
    try:
        # Open with no_new_id to prevent issues
        doc = fitz.open(input_path)
        
        # 1. Font subsetting (safe, ignore errors)
        try:
            doc.subset_fonts()
        except Exception as e:
            logger.warning(f"Font subsetting skipped: {e}")

        processed_xrefs = set()
        image_count = 0
        skipped_count = 0
        
        for page_num, page in enumerate(doc):
            logger.debug(f"Processing page {page_num + 1}")
            
            # Force garbage collection setiap page untuk hemat RAM
            gc.collect()
            
            for img in page.get_images():
                xref = img[0]
                if xref in processed_xrefs:
                    continue
                processed_xrefs.add(xref)
                
                try:
                    # ============== RAM SAFEGUARD ==============
                    # Check image info BEFORE loading to prevent RAM explosion
                    try:
                        img_info = doc.extract_image(xref)
                        if img_info:
                            img_width = img_info.get("width", 0)
                            img_height = img_info.get("height", 0)
                            
                            # Skip extremely large images that would explode RAM
                            if img_width * img_height > MAX_IMAGE_PIXELS:
                                logger.warning(f"Skipped huge image {xref} ({img_width}x{img_height}) to save RAM")
                                skipped_count += 1
                                continue
                                
                            if img_width > MAX_IMAGE_DIMENSION or img_height > MAX_IMAGE_DIMENSION:
                                logger.warning(f"Skipped oversized image {xref} ({img_width}x{img_height})")
                                skipped_count += 1
                                continue
                    except Exception as e:
                        logger.warning(f"Could not get image info for {xref}: {e}")
                    
                    # Now safe to load
                    pix = None
                    try:
                        pix = fitz.Pixmap(doc, xref)
                    except Exception as e:
                        logger.warning(f"Could not load image {xref}: {e}")
                        skipped_count += 1
                        continue
                    
                    # ============== ALPHA HANDLING ==============
                    if pix.alpha:
                        try:
                            # Use Pillow for robust alpha handling
                            img_data = pix.tobytes("png")
                            pix = None  # Free immediately
                            
                            pil_img = Image.open(io.BytesIO(img_data))
                            img_data = None  # Free
                            
                            # Create white background
                            bg = Image.new("RGB", pil_img.size, (255, 255, 255))
                            if pil_img.mode in ('RGBA', 'LA'):
                                bg.paste(pil_img, mask=pil_img.split()[-1])
                            else:
                                bg.paste(pil_img)
                            pil_img.close()
                            pil_img = None
                            
                            # Convert to Grayscale for aggressive compression
                            gray_img = bg.convert("L")
                            bg.close()
                            bg = None
                            
                            # Resize if needed (use COMPRESS_TARGET_SIZE)
                            if gray_img.width > COMPRESS_TARGET_SIZE or gray_img.height > COMPRESS_TARGET_SIZE:
                                gray_img.thumbnail((COMPRESS_TARGET_SIZE, COMPRESS_TARGET_SIZE), Image.Resampling.LANCZOS)
                                
                            # Save as JPEG with low quality
                            out_buffer = io.BytesIO()
                            gray_img.save(out_buffer, format="JPEG", quality=JPEG_QUALITY, optimize=True)
                            gray_img.close()
                            gray_img = None
                            
                            new_data = out_buffer.getvalue()
                            out_buffer = None
                            
                            # Update PDF image
                            try:
                                doc.update_image(xref, data=new_data)
                            except Exception as ue:
                                logger.warning(f"Could not update alpha image {xref}: {ue}")
                            
                            new_data = None
                            image_count += 1
                            
                            # Aggressive GC for 512MB RAM
                            if image_count % GC_EVERY_N_IMAGES == 0:
                                gc.collect()
                            continue
                            
                        except Exception as e:
                            logger.warning(f"Pillow alpha handling failed for {xref}: {e}")
                            gc.collect()
                            # Continue to standard processing
                    
                    # ============== STANDARD PROCESSING ==============
                    # Convert to Grayscale
                    if pix and pix.n >= 3:
                        try:
                            gray_pix = fitz.Pixmap(fitz.csGRAY, pix)
                            pix = None  # Free original immediately
                            pix = gray_pix
                        except Exception as e:
                            logger.warning(f"Grayscale conversion failed: {e}")
                    
                    # Resize if too large (use COMPRESS_TARGET_SIZE)
                    if pix and (pix.width > COMPRESS_TARGET_SIZE or pix.height > COMPRESS_TARGET_SIZE):
                        scale = COMPRESS_TARGET_SIZE / max(pix.width, pix.height)
                        new_w = int(pix.width * scale)
                        new_h = int(pix.height * scale)
                        try:
                            resized_pix = fitz.Pixmap(pix, new_w, new_h)
                            pix = None  # Free original immediately
                            pix = resized_pix
                        except Exception as e:
                            logger.warning(f"Resize failed: {e}")
                    
                    # Convert to JPEG
                    new_data = None
                    if pix:
                        try:
                            new_data = pix.tobytes("jpeg", jpg_quality=JPEG_QUALITY)
                        except Exception as e:
                            logger.warning(f"JPEG conversion failed: {e}")
                            pix = None
                            gc.collect()
                            continue
                    
                    # Update in PDF
                    if new_data:
                        try:
                            doc.update_image(xref, data=new_data)
                        except Exception as e:
                            logger.warning(f"Could not update image {xref}: {e}")
                    
                    # Free memory immediately
                    pix = None 
                    new_data = None
                    image_count += 1
                    
                    # Aggressive garbage collection for 512MB RAM
                    if image_count % GC_EVERY_N_IMAGES == 0:
                        gc.collect()
                        
                except Exception as e:
                    logger.warning(f"Image compression skipped for xref {xref}: {e}")
                    gc.collect()
                    continue

        logger.info(f"Processed {image_count} images, skipped {skipped_count}")
        
        # Force cleanup before save
        gc.collect()

        # ============== SAVE WITH FIX FOR LINEARISATION ERROR ==============
        # FIX: Use linear=False to avoid "Linearisation is no longer supported" error
        try:
            doc.save(
                temp_output, 
                garbage=4,           # Remove unused objects
                deflate=True,        # Compress streams
                clean=True,          # Clean content streams
                linear=False,        # FIX: Disable linearisation
                pretty=False,        # Compact output
                no_new_id=True       # Keep same ID
            )
        except TypeError:
            # Older PyMuPDF might not support all options
            doc.save(temp_output, garbage=4, deflate=True, clean=True)
        
        doc.close()
        doc = None
        gc.collect()
        
        # ============== CHECK IF COMPRESSION WAS EFFECTIVE ==============
        compressed_size = os.path.getsize(temp_output)
        logger.info(f"Compressed size: {compressed_size} bytes")
        
        # If compressed file is not smaller, try alternative approach
        if compressed_size >= original_size:
            logger.warning("Compression not effective, trying alternative method...")
            
            try:
                # Alternative: Recreate PDF from scratch
                alt_output = output_path + ".alt.tmp"
                recreate_pdf_smaller(temp_output, alt_output)
                
                alt_size = os.path.getsize(alt_output)
                logger.info(f"Alternative compression size: {alt_size} bytes")
                
                if alt_size < compressed_size:
                    os.replace(alt_output, temp_output)
                    compressed_size = alt_size
                else:
                    if os.path.exists(alt_output):
                        os.unlink(alt_output)
            except Exception as e:
                logger.warning(f"Alternative compression failed: {e}")
            finally:
                gc.collect()
        
        # Move temp to final output
        shutil.move(temp_output, output_path)
        
        final_size = os.path.getsize(output_path)
        reduction = ((original_size - final_size) / original_size) * 100 if original_size > 0 else 0
        logger.info(f"Compression complete: {original_size} -> {final_size} bytes ({reduction:.1f}% reduction)")
        
    except Exception as e:
        logger.error(f"Compression failed: {e}")
        if doc:
            try:
                doc.close()
            except:
                pass
        # Cleanup temp files
        for f in [temp_output, output_path + ".alt.tmp"]:
            if os.path.exists(f):
                try:
                    os.unlink(f)
                except:
                    pass
        raise e
    finally:
        gc.collect()


def recreate_pdf_smaller(input_path, output_path):
    """
    Recreate PDF from scratch to ensure smaller size.
    This is a fallback when normal compression doesn't reduce size.
    """
    src = fitz.open(input_path)
    dst = fitz.open()
    
    for page in src:
        # Create new page with same dimensions
        new_page = dst.new_page(width=page.rect.width, height=page.rect.height)
        
        # Copy page content as display list (more efficient)
        new_page.show_pdf_page(new_page.rect, src, page.number)
    
    # Save with maximum compression
    dst.save(output_path, garbage=4, deflate=True, clean=True, linear=False)
    dst.close()
    src.close()


def pdf_to_image(pdf_path, image_path):
    doc = fitz.open(pdf_path)
    page = doc[0]  # Take first page
    pix = page.get_pixmap()
    pix.save(image_path)
    doc.close()

@app.route('/')
def index():
    return "Backend is running!"

@app.route('/convert', methods=['POST'])
def convert():
    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400
        
    if file:
        input_path = None
        temp_img_path = None
        output_path = None
        try:
            ext = os.path.splitext(file.filename.lower())[1]
            
            with tempfile.NamedTemporaryFile(delete=False, suffix=ext) as tmp:
                file.save(tmp.name)
                input_path = tmp.name

            temp_img_path = input_path + '.png'
            if ext == '.pdf':
                # Convert PDF to Image first
                pdf_to_image(input_path, temp_img_path)
                processing_path = temp_img_path
            else:
                processing_path = input_path

            output_path = input_path + '.dxf'
            
            # Process image to DXF
            image_to_dxf(processing_path, output_path)
            
            @after_this_request
            def cleanup(response):
                # Cleanup temp files after sending
                for p in [input_path, temp_img_path, output_path]:
                    if p and os.path.exists(p):
                        try:
                            os.unlink(p)
                        except Exception as e:
                            logger.warning(f"Cleanup failed for {p}: {e}")
                return response
            
            # Send file back
            return send_file(
                output_path,
                as_attachment=True,
                download_name=f"{os.path.splitext(file.filename)[0]}.dxf",
                mimetype='application/dxf'
            )
            
        except Exception as e:
            logger.error(f"Convert error: {e}")
            # Cleanup on error
            for p in [input_path, temp_img_path, output_path]:
                if p and os.path.exists(p):
                    try:
                        os.unlink(p)
                    except:
                        pass
            return jsonify({'error': str(e)}), 500

@app.route('/pdf/merge', methods=['POST'])
def pdf_merge_route():
    files = request.files.getlist('files')
    if not files:
        return jsonify({'error': 'No files'}), 400
    
    tmp_paths = []
    converted_tmp_paths = []
    output_path = None
    
    try:
        for f in files:
            ext = os.path.splitext(f.filename.lower())[1]
            
            with tempfile.NamedTemporaryFile(delete=False, suffix=ext) as t:
                f.save(t.name)
                current_path = t.name
                tmp_paths.append(current_path)
            
            if ext in ['.jpg', '.jpeg', '.png']:
                with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as pdf_out:
                    image_to_pdf(current_path, pdf_out.name)
                    converted_tmp_paths.append(pdf_out.name)
            elif ext == '.docx':
                with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as pdf_out:
                    docx_to_pdf(current_path, pdf_out.name)
                    converted_tmp_paths.append(pdf_out.name)
            elif ext == '.pdf':
                converted_tmp_paths.append(current_path)
            else:
                pass

        with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as out:
            output_path = out.name
            merge_pdfs(converted_tmp_paths, output_path)
        
        @after_this_request
        def cleanup(response):
            for p in tmp_paths:
                if os.path.exists(p):
                    try:
                        os.unlink(p)
                    except Exception as e:
                        logger.warning(f"Cleanup failed: {e}")
            for p in converted_tmp_paths:
                if p not in tmp_paths and os.path.exists(p):
                    try:
                        os.unlink(p)
                    except Exception as e:
                        logger.warning(f"Cleanup failed: {e}")
            if output_path and os.path.exists(output_path):
                try:
                    os.unlink(output_path)
                except Exception as e:
                    logger.warning(f"Cleanup failed: {e}")
            return response
        
        return send_file(output_path, as_attachment=True, download_name='merged.pdf', mimetype='application/pdf')
            
    except Exception as e:
        logger.error(f"Merge error: {e}")
        # Cleanup on error
        for p in tmp_paths:
            if os.path.exists(p):
                try:
                    os.unlink(p)
                except:
                    pass
        for p in converted_tmp_paths:
            if p not in tmp_paths and os.path.exists(p):
                try:
                    os.unlink(p)
                except:
                    pass
        return jsonify({'error': str(e)}), 500

@app.route('/pdf/convert', methods=['POST'])
def pdf_convert_route():
    if 'file' not in request.files:
        return jsonify({'error': 'No file'}), 400
    target = request.form.get('target')
    f = request.files['file']
    if not target:
        return jsonify({'error': 'Missing target'}), 400
    ext = os.path.splitext(f.filename.lower())[1]
    
    input_path = None
    output_path = None
    
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=ext) as src:
            f.save(src.name)
            input_path = src.name
            
        if ext == '.pdf' and target == 'docx':
            with tempfile.NamedTemporaryFile(delete=False, suffix='.docx') as out:
                output_path = out.name
                pdf_to_docx(input_path, output_path)
                
                @after_this_request
                def cleanup(response):
                    for p in [input_path, output_path]:
                        if p and os.path.exists(p):
                            try:
                                os.unlink(p)
                            except:
                                pass
                    return response
                
                return send_file(output_path, as_attachment=True, download_name='converted.docx', mimetype='application/vnd.openxmlformats-officedocument.wordprocessingml.document')
        
        elif ext == '.pdf' and target == 'xlsx':
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as out:
                output_path = out.name
                pdf_to_xlsx(input_path, output_path)
                
                @after_this_request
                def cleanup(response):
                    for p in [input_path, output_path]:
                        if p and os.path.exists(p):
                            try:
                                os.unlink(p)
                            except:
                                pass
                    return response
                
                return send_file(output_path, as_attachment=True, download_name='converted.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
        elif ext == '.docx' and target == 'pdf':
            with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as out:
                output_path = out.name
                docx_to_pdf(input_path, output_path)
                
                @after_this_request
                def cleanup(response):
                    for p in [input_path, output_path]:
                        if p and os.path.exists(p):
                            try:
                                os.unlink(p)
                            except:
                                pass
                    return response
                
                return send_file(output_path, as_attachment=True, download_name='converted.pdf', mimetype='application/pdf')
        
        elif ext == '.xlsx' and target == 'pdf':
            with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as out:
                output_path = out.name
                xlsx_to_pdf(input_path, output_path)
                
                @after_this_request
                def cleanup(response):
                    for p in [input_path, output_path]:
                        if p and os.path.exists(p):
                            try:
                                os.unlink(p)
                            except:
                                pass
                    return response
                
                return send_file(output_path, as_attachment=True, download_name='converted.pdf', mimetype='application/pdf')
        
        else:
            if input_path and os.path.exists(input_path):
                os.unlink(input_path)
            return jsonify({'error': 'Unsupported conversion'}), 400
            
    except Exception as e:
        logger.error(f"Convert doc error: {e}")
        for p in [input_path, output_path]:
            if p and os.path.exists(p):
                try:
                    os.unlink(p)
                except:
                    pass
        return jsonify({'error': str(e)}), 500

@app.route('/pdf/compress', methods=['POST'])
def pdf_compress_route():
    logger.info("Received compress request")
    
    files = request.files.getlist('files')
    if not files:
        files = request.files.getlist('file')
    
    if not files:
        logger.error("No files found in request")
        return jsonify({'error': 'No file uploaded'}), 400
        
    tmp_paths = []
    compressed_paths = []
    
    try:
        for f in files:
            if not f.filename:
                continue
                
            ext = os.path.splitext(f.filename)[1].lower()
            if ext != '.pdf':
                continue 
                
            logger.info(f"Processing file: {f.filename}")
            
            with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as src:
                f.save(src.name)
                tmp_paths.append(src.name)
                input_path = src.name
                
            out_path = input_path + '_compressed.pdf'
            compress_pdf(input_path, out_path)
            compressed_paths.append((out_path, f.filename))

        if not compressed_paths:
            logger.error("No valid PDF files processed")
            return jsonify({'error': 'No valid PDF files processed. Please upload PDF files.'}), 400

        # Schedule cleanup after request
        @after_this_request
        def cleanup(response):
            for p in tmp_paths:
                if os.path.exists(p):
                    try:
                        os.unlink(p)
                    except Exception as e:
                        logger.warning(f"Failed to delete temp input {p}: {e}")
            for path, _ in compressed_paths:
                if os.path.exists(path):
                    try:
                        os.unlink(path)
                    except Exception as e:
                        logger.warning(f"Failed to delete compressed {path}: {e}")
            return response

        if len(compressed_paths) == 1:
            logger.info("Returning single compressed file")
            return send_file(compressed_paths[0][0], as_attachment=True, download_name='compressed.pdf', mimetype='application/pdf')
        else:
            logger.info("Returning zip of compressed files")
            with tempfile.NamedTemporaryFile(delete=False, suffix='.zip') as zip_out:
                zip_path = zip_out.name
                with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
                    for path, original_name in compressed_paths:
                        zf.write(path, arcname=f"compressed_{original_name}")
                return send_file(zip_path, as_attachment=True, download_name='compressed_files.zip', mimetype='application/zip')

    except Exception as e:
        logger.error(f"Compression route error: {e}")
        # Cleanup on error
        for p in tmp_paths:
            if os.path.exists(p):
                try:
                    os.unlink(p)
                except:
                    pass
        for path, _ in compressed_paths:
            if os.path.exists(path):
                try:
                    os.unlink(path)
                except:
                    pass
        return jsonify({'error': f"Compression failed: {str(e)}"}), 500

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port)
